from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QComboBox,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QDateEdit, QFileDialog, QMessageBox
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QColor
import pandas as pd
import openpyxl
from ..utils.ui_helper import UIHelper

class ReportsManager:
    def __init__(self, database):
        self.database = database

    def generate_financial_report(self, start_date, end_date):
        """توليد تقرير الإيرادات والمصروفات"""
        try:
            # التأكد من وجود اتصال نشط بقاعدة البيانات
            self.database.ensure_connection()
            
            self.database.cursor.execute("""
                SELECT entry_type,
                       category,
                       SUM(amount) as total,
                       COUNT(*) as count
                FROM financial_entries
                WHERE date BETWEEN ? AND ?
                GROUP BY entry_type, category
                ORDER BY entry_type, category
            """, (start_date, end_date))
            
            results = self.database.cursor.fetchall()
            
            summary = {
                "إيراد": {"total": 0, "details": {}},
                "مصروف": {"total": 0, "details": {}}
            }
            
            for entry_type, category, total, count in results:
                summary[entry_type]["total"] += total
                summary[entry_type]["details"][category] = {
                    "total": total,
                    "count": count
                }
            
            return True, results, summary, None
            
        except Exception as e:
            print(f"Error in generate_financial_report: {str(e)}")
            return False, [], {}, str(e)

    def generate_installments_report(self, start_date, end_date):
        """توليد تقرير الأقساط"""
        try:
            # التأكد من وجود اتصال نشط بقاعدة البيانات
            self.database.ensure_connection()
            
            self.database.cursor.execute("""
                SELECT i.id,
                       c.brand || ' ' || c.model as car_name,
                       cl.name as client_name,
                       i.total_amount,
                       i.paid_amount,
                       i.remaining_amount,
                       i.installment_count,
                       i.next_payment_date,
                       i.status
                FROM installments i
                JOIN cars c ON i.car_id = c.id
                JOIN clients cl ON i.client_id = cl.id
                WHERE i.start_date BETWEEN ? AND ?
                ORDER BY i.status, i.next_payment_date
            """, (start_date, end_date))
            
            results = self.database.cursor.fetchall()
            
            summary = {
                "total_count": len(results),
                "total_amount": sum(row[3] for row in results),
                "total_paid": sum(row[4] for row in results),
                "total_remaining": sum(row[5] for row in results),
                "status": {
                    "جاري": len([r for r in results if r[8] == "جاري"]),
                    "متأخر": len([r for r in results if r[8] == "متأخر"]),
                    "منتهي": len([r for r in results if r[8] == "منتهي"])
                }
            }
            
            return True, results, summary, None
            
        except Exception as e:
            print(f"Error in generate_installments_report: {str(e)}")
            return False, [], {}, str(e)

    def generate_sales_report(self, start_date, end_date):
        """توليد تقرير المبيعات"""
        try:
            # التأكد من وجود اتصال نشط بقاعدة البيانات
            self.database.ensure_connection()
            
            self.database.cursor.execute("""
                SELECT i.invoice_number,
                       c.brand || ' ' || c.model AS car_name,
                       cl.name AS client_name,
                       i.invoice_date,
                       i.total_amount,
                       i.payment_method,
                       i.payment_status
                FROM invoices i
                JOIN cars c ON i.car_id = c.id
                JOIN clients cl ON i.client_id = cl.id
                WHERE i.invoice_date BETWEEN ? AND ?
                ORDER BY i.invoice_date DESC
            """, (start_date, end_date))
            
            results = self.database.cursor.fetchall()
            
            summary = {
                "total_count": len(results),
                "total_amount": sum(row[4] for row in results),
                "payment_method": {
                    "نقدي": {
                        "count": len([r for r in results if r[5] == "نقدي"]),
                        "total": sum(r[4] for r in results if r[5] == "نقدي")
                    },
                    "تقسيط": {
                        "count": len([r for r in results if r[5] == "تقسيط"]),
                        "total": sum(r[4] for r in results if r[5] == "تقسيط")
                    }
                }
            }
            
            return True, results, summary, None
            
        except Exception as e:
            print(f"Error in generate_sales_report: {str(e)}")
            return False, [], {}, str(e)

    def generate_clients_report(self, start_date, end_date):
        """توليد تقرير العملاء"""
        try:
            # التأكد من وجود اتصال نشط بقاعدة البيانات
            self.database.ensure_connection()
            
            self.database.cursor.execute("""
                SELECT cl.name,
                       cl.phone,
                       COUNT(DISTINCT i.id) as invoices_count,
                       SUM(i.total_amount) as total_amount,
                       COUNT(DISTINCT inst.id) as installments_count,
                       SUM(inst.remaining_amount) as remaining_amount
                FROM clients cl
                LEFT JOIN invoices i ON cl.id = i.client_id
                    AND i.invoice_date BETWEEN ? AND ?
                LEFT JOIN installments inst ON cl.id = inst.client_id
                    AND inst.start_date BETWEEN ? AND ?
                GROUP BY cl.id
                ORDER BY total_amount DESC NULLS LAST
            """, (start_date, end_date, start_date, end_date))
            
            results = self.database.cursor.fetchall()
            
            summary = {
                "total_clients": len(results),
                "total_sales": sum(row[3] or 0 for row in results),
                "total_remaining": sum(row[5] or 0 for row in results),
                "with_installments": len([r for r in results if r[4] and r[4] > 0])
            }
            
            return True, results, summary, None
            
        except Exception as e:
            print(f"Error in generate_clients_report: {str(e)}")
            return False, [], {}, str(e)

    def export_to_excel(self, data, headers, filename):
        """تصدير البيانات إلى ملف Excel"""
        try:
            df = pd.DataFrame(data, columns=headers)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='البيانات', index=False)
                
                # تنسيق العناوين
                worksheet = writer.sheets['البيانات']
                for col in range(len(headers)):
                    cell = worksheet.cell(row=1, column=col+1)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color='f8f9fa',
                        end_color='f8f9fa',
                        fill_type='solid'
                    )
                    
                    # تعيين العرض المناسب للعمود
                    worksheet.column_dimensions[
                        openpyxl.utils.get_column_letter(col+1)
                    ].width = 15
            
            return True, None
            
        except Exception as e:
            print(f"Error in export_to_excel: {str(e)}")
            return False, str(e)
